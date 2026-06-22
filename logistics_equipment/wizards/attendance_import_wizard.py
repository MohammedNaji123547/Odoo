import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class LogisticsAttendanceImportWizard(models.TransientModel):
    _name = 'logistics.attendance.import.wizard'
    _description = 'Attendance Excel Import / Export Wizard'

    # ── Export filters ────────────────────────────────────────────────────────
    date_from = fields.Date(string='From Date', default=fields.Date.today)
    date_to = fields.Date(string='To Date', default=fields.Date.today)
    request_id = fields.Many2one(
        'logistics.request', string='Request (optional)',
        domain=[('state', '=', 'attendance_confirmation')]
    )

    # ── Import ────────────────────────────────────────────────────────────────
    import_file = fields.Binary(string='Upload Filled Template')
    import_filename = fields.Char()

    # ── Result ────────────────────────────────────────────────────────────────
    state = fields.Selection([
        ('draft', 'Draft'),
        ('done',  'Done'),
    ], default='draft')
    result_message = fields.Text(string='Result', readonly=True)

    # ─────────────────────────────────────────────────────────────────────────
    def action_download_template(self):
        """Generate an Excel file of pending attendance records for the user to fill."""
        if not openpyxl:
            raise UserError(_('openpyxl is not available. Contact your system administrator.'))

        domain = [('status', '=', 'pending')]
        if self.date_from:
            domain.append(('attendance_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('attendance_date', '<=', self.date_to))
        if self.request_id:
            domain.append(('request_id', '=', self.request_id.id))

        records = self.env['logistics.attendance'].search(
            domain, order='attendance_date asc, equipment_id asc'
        )
        if not records:
            raise UserError(_(
                'No pending attendance records found for the selected criteria.'
            ))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Attendance'

        # ── Styles ────────────────────────────────────────────────────────────
        hdr_font   = Font(bold=True, color='FFFFFF', size=11)
        hdr_fill   = PatternFill('solid', fgColor='1F4E79')
        lock_fill  = PatternFill('solid', fgColor='DCE6F1')  # read-only cols
        edit_fill  = PatternFill('solid', fgColor='FFFFFF')
        center     = Alignment(horizontal='center', vertical='center')
        wrap       = Alignment(wrap_text=True, vertical='center')
        thin_side  = openpyxl.styles.Side(style='thin', color='B0B0B0')
        thin_border = openpyxl.styles.Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        # ── Row 1: instructions ───────────────────────────────────────────────
        ws.merge_cells('A1:J1')
        ws['A1'] = (
            'INSTRUCTIONS: Columns A–F are read-only references. '
            'Fill in columns G (Status), H (Actual Qty), I (Working Hours), J (Remarks) only. '
            'Valid Status values: present | partial | absent'
        )
        ws['A1'].font = Font(bold=True, color='5C3317', size=10)
        ws['A1'].fill = PatternFill('solid', fgColor='FFF3CD')
        ws['A1'].alignment = wrap
        ws.row_dimensions[1].height = 30

        # ── Row 2: column headers ─────────────────────────────────────────────
        headers = [
            'ID *',
            'Date',
            'Equipment',
            'Supplier',
            'Request',
            'Planned Qty',
            'Status',
            'Actual Qty',
            'Working Hours',
            'Remarks',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin_border
        ws.row_dimensions[2].height = 20

        # Freeze panes so headers stay visible when scrolling
        ws.freeze_panes = 'G3'

        # ── Status dropdown validation ─────────────────────────────────────────
        dv = DataValidation(
            type='list',
            formula1='"present,partial,absent"',
            showDropDown=False,
            error='Please choose: present, partial, or absent',
            errorTitle='Invalid Value',
            showErrorMessage=True,
        )
        ws.add_data_validation(dv)

        # ── Data rows ─────────────────────────────────────────────────────────
        for row_idx, rec in enumerate(records, 3):
            values = [
                rec.id,
                rec.attendance_date,
                rec.equipment_id.display_name or '',
                rec.supplier_id.name or '',
                rec.request_id.name or '',
                rec.planned_qty,
                'present',          # suggested default
                rec.planned_qty,    # suggested default
                10.0,               # suggested default
                '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(vertical='center')
                cell.border = thin_border
                if col <= 6:
                    # read-only reference columns — shade them
                    cell.fill = lock_fill
                    cell.font = Font(color='444444')
                else:
                    cell.fill = edit_fill

            # Apply dropdown to Status cell (column G = 7)
            dv.add(ws.cell(row=row_idx, column=7))

        # ── Column widths ─────────────────────────────────────────────────────
        col_widths = [8, 13, 30, 25, 16, 12, 12, 12, 15, 35]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Save & return download ─────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        file_bytes = base64.b64encode(output.getvalue())

        date_from = str(self.date_from or 'all')
        date_to   = str(self.date_to   or 'all')
        filename  = f'attendance_{date_from}_to_{date_to}.xlsx'

        attachment = self.env['ir.attachment'].create({
            'name':     filename,
            'type':     'binary',
            'datas':    file_bytes,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    # ─────────────────────────────────────────────────────────────────────────
    def action_import(self):
        """Parse uploaded Excel and bulk-update attendance records."""
        if not self.import_file:
            raise UserError(_('Please upload an Excel file first.'))
        if not openpyxl:
            raise UserError(_('openpyxl is not available. Contact your system administrator.'))

        file_data = base64.b64decode(self.import_file)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_('Could not read the Excel file: %s') % str(e))

        ws = wb.active
        VALID_STATUSES = {'present', 'partial', 'absent'}

        updated = 0
        skipped = 0
        error_lines = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
            if not row or row[0] is None:
                continue  # blank row

            raw_id       = row[0]
            raw_status   = row[6]
            raw_actual   = row[7]
            raw_hours    = row[8]
            raw_remarks  = row[9]

            # Validate ID
            try:
                rec_id = int(raw_id)
            except (ValueError, TypeError):
                error_lines.append(f'Row {row_idx}: invalid ID "{raw_id}" — skipped')
                skipped += 1
                continue

            # Validate status
            status = str(raw_status).strip().lower() if raw_status else ''
            if status not in VALID_STATUSES:
                error_lines.append(
                    f'Row {row_idx}: invalid status "{raw_status}" '
                    f'(expected present/partial/absent) — skipped'
                )
                skipped += 1
                continue

            # Find the record
            attendance = self.env['logistics.attendance'].browse(rec_id)
            if not attendance.exists():
                error_lines.append(f'Row {row_idx}: ID {rec_id} not found — skipped')
                skipped += 1
                continue

            # Build write vals
            vals = {
                'status':  status,
                'remarks': str(raw_remarks) if raw_remarks else '',
            }
            if status != 'absent':
                try:
                    vals['actual_qty']     = float(raw_actual or 0)
                    vals['working_hours']  = float(raw_hours  or 10)
                except (ValueError, TypeError):
                    error_lines.append(
                        f'Row {row_idx}: non-numeric Actual Qty or Working Hours — '
                        f'defaulted to 0 / 10'
                    )
                    vals.setdefault('actual_qty',    0.0)
                    vals.setdefault('working_hours', 10.0)

            attendance.write(vals)
            updated += 1

        summary = [f'✓ {updated} record(s) updated.']
        if skipped:
            summary.append(f'⚠ {skipped} row(s) skipped.')
        summary.extend(error_lines[:30])  # cap at 30 error lines

        self.write({
            'state':          'done',
            'result_message': '\n'.join(summary),
        })
        # Re-open wizard to show result
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }
